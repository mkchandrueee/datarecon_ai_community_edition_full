"""Unit tests — ReportingService (Module 18)."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from io import BytesIO

import pandas as pd
import pytest

from datarecon.application.services.reporting_service import (
    ReportingError,
    ReportingService,
    ReportPayload,
    ReportSection,
    sanitize_export_name,
)
from datarecon.domain.enums import ReportFormat


@pytest.fixture
def service() -> ReportingService:
    return ReportingService()


@pytest.fixture
def payload() -> ReportPayload:
    return ReportPayload(
        title="Record Count Validation",
        summary={"source_count": 100, "target_count": 95, "status": "FAIL"},
        sections=(
            ReportSection(
                "Group Breakdown",
                pd.DataFrame({"region": ["east", "west"], "difference": [-3, -2]}),
            ),
        ),
    )


def test_json_export_roundtrips(service: ReportingService, payload: ReportPayload) -> None:
    raw = service.export(payload, ReportFormat.JSON)
    doc = json.loads(raw)
    assert doc["title"] == "Record Count Validation"
    assert doc["summary"]["source_count"] == 100
    assert doc["sections"]["Group Breakdown"][0]["region"] == "east"


def test_csv_export_single_section(service: ReportingService, payload: ReportPayload) -> None:
    raw = service.export(payload, ReportFormat.CSV)
    text = raw.decode("utf-8")
    assert "region,difference" in text
    assert "east,-3" in text


def test_csv_export_stacks_multiple_sections_under_banners(service: ReportingService) -> None:
    payload = ReportPayload(
        title="x",
        summary={"rows": 3},
        sections=(
            ReportSection("A", pd.DataFrame({"a": [1]})),
            ReportSection("B", pd.DataFrame({"b": [2]})),
        ),
    )
    text = service.export(payload, ReportFormat.CSV).decode("utf-8")
    assert "# Summary" in text
    assert "rows,3" in text
    assert "# A" in text
    assert "a\n1" in text
    assert "# B" in text
    assert "b\n2" in text


def test_csv_export_falls_back_to_summary_when_no_sections(service: ReportingService) -> None:
    payload = ReportPayload(title="x", summary={"source_count": 10, "target_count": 10})
    text = service.export(payload, ReportFormat.CSV).decode("utf-8")
    assert "metric,value" in text
    assert "source_count,10" in text
    assert "target_count,10" in text


def test_csv_line_endings_are_lf_on_every_platform(service: ReportingService) -> None:
    """Pinned so a report built on Windows matches one built on Linux, and the
    banner lines never mix with pandas' CRLF inside a single file."""
    payload = ReportPayload(
        title="x",
        summary={"rows": 1},
        sections=(
            ReportSection("A", pd.DataFrame({"a": [1]})),
            ReportSection("B", pd.DataFrame({"b": [2]})),
        ),
    )
    for shape in (payload, ReportPayload(title="x", summary={"rows": 1})):
        assert b"\r\n" not in service.export(shape, ReportFormat.CSV)


def test_csv_export_of_empty_section_still_emits_header(service: ReportingService) -> None:
    payload = ReportPayload(
        title="x",
        summary={},
        sections=(ReportSection("Mismatches", pd.DataFrame(columns=["id", "email"])),),
    )
    text = service.export(payload, ReportFormat.CSV).decode("utf-8")
    assert "id,email" in text


def test_excel_export_produces_nonempty_workbook(
    service: ReportingService, payload: ReportPayload
) -> None:
    raw = service.export(payload, ReportFormat.EXCEL)
    assert raw[:2] == b"PK"  # xlsx is a zip archive
    assert len(raw) > 100


def test_excel_export_handles_many_sections_and_long_titles(service: ReportingService) -> None:
    sections = tuple(
        ReportSection(f"Section With A Very Long Title Number {i}", pd.DataFrame({"x": [i]}))
        for i in range(5)
    )
    payload = ReportPayload(title="stress", summary={"n": 5}, sections=sections)
    raw = service.export(payload, ReportFormat.EXCEL)
    assert raw[:2] == b"PK"


def test_pdf_export_produces_pdf_bytes(service: ReportingService, payload: ReportPayload) -> None:
    raw = service.export(payload, ReportFormat.PDF)
    assert raw[:4] == b"%PDF"


def test_pdf_export_handles_empty_sections(service: ReportingService) -> None:
    payload = ReportPayload(title="empty", summary={"n": 0}, sections=())
    raw = service.export(payload, ReportFormat.PDF)
    assert raw[:4] == b"%PDF"


def test_unsupported_format_raises(service: ReportingService, payload: ReportPayload) -> None:
    with pytest.raises(ReportingError, match="Unsupported report format"):
        service.export(payload, "xml")  # type: ignore[arg-type]


def test_content_type_and_extension_mapping(service: ReportingService) -> None:
    assert service.file_extension(ReportFormat.EXCEL) == "xlsx"
    assert service.file_extension(ReportFormat.CSV) == "csv"
    assert service.content_type(ReportFormat.JSON) == "application/json"


# ---------- Excel and timezone-aware timestamps ----------


def test_excel_export_handles_tz_aware_datetime_column(service: ReportingService) -> None:
    """xlsx has no tz-aware datetime type; xlsxwriter raises rather than guess."""
    table = pd.DataFrame(
        {"run": ["a", "b"], "started_at": pd.to_datetime(["2026-01-01", "2026-01-02"], utc=True)}
    )
    payload = ReportPayload(title="x", summary={}, sections=(ReportSection("Runs", table),))

    raw = service.export(payload, ReportFormat.EXCEL)

    assert raw[:2] == b"PK"


def test_excel_export_handles_object_column_of_tz_aware_values(
    service: ReportingService,
) -> None:
    """Datetimes mixed with None stay object-dtype, so .dt is unavailable."""
    table = pd.DataFrame(
        {"suite": ["a", "b"], "last_run": [datetime(2026, 1, 1, tzinfo=UTC), None]}
    )
    payload = ReportPayload(title="x", summary={}, sections=(ReportSection("Suites", table),))

    raw = service.export(payload, ReportFormat.EXCEL)

    assert raw[:2] == b"PK"


def test_excel_export_handles_tz_aware_value_in_summary(service: ReportingService) -> None:
    payload = ReportPayload(title="x", summary={"generated": datetime(2026, 1, 1, tzinfo=UTC)})

    raw = service.export(payload, ReportFormat.EXCEL)

    assert raw[:2] == b"PK"


def test_excel_safe_leaves_naive_frames_untouched(service: ReportingService) -> None:
    table = pd.DataFrame({"a": [1], "when": pd.to_datetime(["2026-01-01"])})
    assert service._excel_safe(table) is table


# ---------- batched extracts (ADR-0013) ----------


def test_small_frame_is_a_single_unnumbered_batch(service: ReportingService) -> None:
    """Ordinary downloads keep their plain name — no _1 suffix appears."""
    df = pd.DataFrame({"a": range(10)})

    batches = service.batch_frame("Mismatches", df, batch_rows=50)

    assert len(batches) == 1
    assert batches[0].name == "Mismatches"
    assert batches[0].is_only_batch
    assert (batches[0].first_row, batches[0].last_row) == (1, 10)


def test_frame_exactly_at_threshold_is_not_split(service: ReportingService) -> None:
    batches = service.batch_frame("X", pd.DataFrame({"a": range(50)}), batch_rows=50)

    assert len(batches) == 1
    assert batches[0].name == "X"


def test_large_frame_splits_into_numbered_batches(service: ReportingService) -> None:
    df = pd.DataFrame({"a": range(120)})

    batches = service.batch_frame("DV_CUSTOMER_MASTER_PASS", df, batch_rows=50)

    assert [b.name for b in batches] == [
        "DV_CUSTOMER_MASTER_PASS_1",
        "DV_CUSTOMER_MASTER_PASS_2",
        "DV_CUSTOMER_MASTER_PASS_3",
    ]
    assert [b.total for b in batches] == [3, 3, 3]
    assert [(b.first_row, b.last_row) for b in batches] == [(1, 50), (51, 100), (101, 120)]


def test_batches_partition_the_frame_without_loss_or_overlap(
    service: ReportingService,
) -> None:
    """Batching is about file size, never about dropping rows."""
    df = pd.DataFrame({"a": range(1_000)})

    batches = service.batch_frame("X", df, batch_rows=137)

    rebuilt = pd.concat([b.dataframe for b in batches], ignore_index=True)
    pd.testing.assert_frame_equal(rebuilt, df)
    assert sum(len(b.dataframe) for b in batches) == len(df)


def test_empty_frame_still_yields_one_downloadable_batch(service: ReportingService) -> None:
    batches = service.batch_frame("Mismatches", pd.DataFrame({"a": []}), batch_rows=50)

    assert len(batches) == 1
    assert batches[0].row_range_label == "no rows"


def test_batch_rows_below_one_is_rejected(service: ReportingService) -> None:
    with pytest.raises(ReportingError, match="at least 1"):
        service.batch_frame("X", pd.DataFrame({"a": [1]}), batch_rows=0)


def test_row_range_label_is_thousands_separated(service: ReportingService) -> None:
    batches = service.batch_frame("X", pd.DataFrame({"a": range(3_000)}), batch_rows=2_000)

    assert batches[0].row_range_label == "rows 1-2,000"
    assert batches[1].row_range_label == "rows 2,001-3,000"


def test_oversized_section_spans_numbered_excel_sheets(service: ReportingService) -> None:
    """xlsx caps a sheet at ~1M rows; the export must not fail on that."""
    from openpyxl import load_workbook

    import datarecon.application.services.reporting_service as module

    original = module.EXCEL_MAX_ROWS_PER_SHEET
    module.EXCEL_MAX_ROWS_PER_SHEET = 10
    try:
        payload = ReportPayload(
            title="x",
            summary={},
            sections=(ReportSection("Mismatches", pd.DataFrame({"a": range(25)})),),
        )
        raw = service.export(payload, ReportFormat.EXCEL)
    finally:
        module.EXCEL_MAX_ROWS_PER_SHEET = original

    workbook = load_workbook(BytesIO(raw))
    assert workbook.sheetnames == ["Summary", "Mismatches_1", "Mismatches_2", "Mismatches_3"]


# ---------- export filenames ----------


def test_sanitize_export_name_strips_filesystem_unsafe_characters() -> None:
    assert sanitize_export_name("DV_CUSTOMER MASTER / v2") == "DV_CUSTOMER_MASTER_v2"


def test_sanitize_export_name_falls_back_when_nothing_survives() -> None:
    assert sanitize_export_name("///") == "export"


def test_sanitize_export_name_leaves_a_clean_name_alone() -> None:
    assert sanitize_export_name("DV_CUSTOMER_MASTER_PASS_1") == "DV_CUSTOMER_MASTER_PASS_1"
